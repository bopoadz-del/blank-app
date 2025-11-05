"""
Report generation service for creating PDF, Excel, and CSV reports.
"""
from sqlalchemy.orm import Session
from typing import Optional, Dict, Any, List
from datetime import datetime
import io
import csv
import logging
from pathlib import Path

from app.models.notifications import Report, ReportFormat, ReportType
from app.models.auth import User
from app.models.chat import Project, Conversation, Message

logger = logging.getLogger(__name__)


class ReportService:
    """Service for generating reports."""

    @staticmethod
    def create_report(
        db: Session,
        user_id: int,
        title: str,
        report_type: ReportType,
        report_format: ReportFormat,
        file_path: str,
        file_size: int,
        metadata: Optional[Dict[str, Any]] = None
    ) -> Report:
        """Create a report record."""
        report = Report(
            user_id=user_id,
            title=title,
            type=report_type,
            format=report_format,
            file_path=file_path,
            file_size=file_size,
            metadata=metadata or {}
        )
        db.add(report)
        db.commit()
        db.refresh(report)
        logger.info(f"Created report {report.id} for user {user_id}")
        return report

    @staticmethod
    def get_user_reports(
        db: Session,
        user_id: int,
        report_type: Optional[ReportType] = None,
        limit: int = 50
    ) -> List[Report]:
        """Get reports for a user."""
        query = db.query(Report).filter(Report.user_id == user_id)

        if report_type:
            query = query.filter(Report.type == report_type)

        return query.order_by(Report.created_at.desc()).limit(limit).all()

    @staticmethod
    def delete_report(db: Session, report_id: int, user_id: int) -> bool:
        """Delete a report."""
        report = db.query(Report).filter(
            Report.id == report_id,
            Report.user_id == user_id
        ).first()

        if report:
            # Delete the file
            try:
                file_path = Path(report.file_path)
                if file_path.exists():
                    file_path.unlink()
            except Exception as e:
                logger.error(f"Error deleting report file: {str(e)}")

            db.delete(report)
            db.commit()
            logger.info(f"Deleted report {report_id} for user {user_id}")
            return True

        return False

    @staticmethod
    def generate_conversation_csv(
        db: Session,
        conversation_id: int,
        user_id: int
    ) -> bytes:
        """Generate CSV report for a conversation."""
        conversation = db.query(Conversation).filter(
            Conversation.id == conversation_id,
            Conversation.user_id == user_id
        ).first()

        if not conversation:
            raise ValueError("Conversation not found")

        output = io.StringIO()
        writer = csv.writer(output)

        # Header
        writer.writerow(['Timestamp', 'Role', 'Content', 'Message ID'])

        # Messages
        for message in conversation.messages:
            writer.writerow([
                message.created_at.isoformat(),
                message.role,
                message.content,
                message.id
            ])

        return output.getvalue().encode('utf-8')

    @staticmethod
    def generate_project_csv(
        db: Session,
        project_id: int,
        user_id: int
    ) -> bytes:
        """Generate CSV report for a project."""
        project = db.query(Project).filter(
            Project.id == project_id,
            Project.user_id == user_id
        ).first()

        if not project:
            raise ValueError("Project not found")

        output = io.StringIO()
        writer = csv.writer(output)

        # Header
        writer.writerow(['Conversation', 'Message Count', 'Created At', 'Last Updated'])

        # Conversations
        for conv in project.conversations:
            writer.writerow([
                conv.title,
                len(conv.messages),
                conv.created_at.isoformat(),
                (conv.updated_at or conv.created_at).isoformat()
            ])

        return output.getvalue().encode('utf-8')

    @staticmethod
    def generate_conversation_excel(
        db: Session,
        conversation_id: int,
        user_id: int
    ) -> bytes:
        """Generate Excel report for a conversation."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            raise ImportError("openpyxl is required for Excel generation")

        conversation = db.query(Conversation).filter(
            Conversation.id == conversation_id,
            Conversation.user_id == user_id
        ).first()

        if not conversation:
            raise ValueError("Conversation not found")

        wb = Workbook()
        ws = wb.active
        ws.title = "Conversation"

        # Header
        headers = ['Timestamp', 'Role', 'Content', 'Message ID']
        ws.append(headers)

        # Style header
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill

        # Messages
        for message in conversation.messages:
            ws.append([
                message.created_at,
                message.role,
                message.content,
                message.id
            ])

        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 60
        ws.column_dimensions['D'].width = 10

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    @staticmethod
    def generate_project_excel(
        db: Session,
        project_id: int,
        user_id: int
    ) -> bytes:
        """Generate Excel report for a project."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            raise ImportError("openpyxl is required for Excel generation")

        project = db.query(Project).filter(
            Project.id == project_id,
            Project.user_id == user_id
        ).first()

        if not project:
            raise ValueError("Project not found")

        wb = Workbook()
        ws = wb.active
        ws.title = "Project Summary"

        # Header
        headers = ['Conversation', 'Message Count', 'Created At', 'Last Updated']
        ws.append(headers)

        # Style header
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill

        # Conversations
        for conv in project.conversations:
            ws.append([
                conv.title,
                len(conv.messages),
                conv.created_at,
                conv.updated_at or conv.created_at
            ])

        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    @staticmethod
    def generate_conversation_pdf(
        db: Session,
        conversation_id: int,
        user_id: int
    ) -> bytes:
        """Generate PDF report for a conversation."""
        try:
            from reportlab.lib.pagesizes import letter
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
            from reportlab.lib import colors
        except ImportError:
            raise ImportError("reportlab is required for PDF generation")

        conversation = db.query(Conversation).filter(
            Conversation.id == conversation_id,
            Conversation.user_id == user_id
        ).first()

        if not conversation:
            raise ValueError("Conversation not found")

        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=letter)
        story = []
        styles = getSampleStyleSheet()

        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#4A90E2')
        )
        title = Paragraph(f"Conversation: {conversation.title}", title_style)
        story.append(title)
        story.append(Spacer(1, 0.3 * inch))

        # Metadata
        metadata_text = f"Created: {conversation.created_at.strftime('%Y-%m-%d %H:%M')}<br/>" \
                       f"Messages: {len(conversation.messages)}"
        metadata = Paragraph(metadata_text, styles['Normal'])
        story.append(metadata)
        story.append(Spacer(1, 0.3 * inch))

        # Messages
        for message in conversation.messages:
            # Role header
            role_style = ParagraphStyle(
                'Role',
                parent=styles['Heading3'],
                textColor=colors.HexColor('#4A90E2') if message.role == 'user' else colors.HexColor('#50C878')
            )
            role = Paragraph(f"{message.role.upper()}", role_style)
            story.append(role)

            # Message content
            content = Paragraph(message.content, styles['Normal'])
            story.append(content)
            story.append(Spacer(1, 0.2 * inch))

        doc.build(story)
        return output.getvalue()

    @staticmethod
    def generate_project_pdf(
        db: Session,
        project_id: int,
        user_id: int
    ) -> bytes:
        """Generate PDF report for a project."""
        try:
            from reportlab.lib.pagesizes import letter
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
            from reportlab.lib import colors
        except ImportError:
            raise ImportError("reportlab is required for PDF generation")

        project = db.query(Project).filter(
            Project.id == project_id,
            Project.user_id == user_id
        ).first()

        if not project:
            raise ValueError("Project not found")

        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=letter)
        story = []
        styles = getSampleStyleSheet()

        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#4A90E2')
        )
        title = Paragraph(f"Project: {project.name}", title_style)
        story.append(title)
        story.append(Spacer(1, 0.3 * inch))

        # Description
        if project.description:
            desc = Paragraph(f"<b>Description:</b> {project.description}", styles['Normal'])
            story.append(desc)
            story.append(Spacer(1, 0.2 * inch))

        # Metadata
        metadata_text = f"Created: {project.created_at.strftime('%Y-%m-%d %H:%M')}<br/>" \
                       f"Conversations: {len(project.conversations)}"
        metadata = Paragraph(metadata_text, styles['Normal'])
        story.append(metadata)
        story.append(Spacer(1, 0.3 * inch))

        # Conversations table
        table_data = [['Conversation', 'Messages', 'Created', 'Updated']]
        for conv in project.conversations:
            table_data.append([
                conv.title,
                str(len(conv.messages)),
                conv.created_at.strftime('%Y-%m-%d'),
                (conv.updated_at or conv.created_at).strftime('%Y-%m-%d')
            ])

        table = Table(table_data, colWidths=[3*inch, 1*inch, 1.5*inch, 1.5*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4A90E2')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(table)

        doc.build(story)
        return output.getvalue()


report_service = ReportService()
